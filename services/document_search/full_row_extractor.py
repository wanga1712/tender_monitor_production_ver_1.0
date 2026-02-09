"""
MODULE: services.document_search.full_row_extractor
RESPONSIBILITY: Extract full row data with context from Excel files.
ALLOWED: openpyxl, xlrd, logging, excel_utils.
FORBIDDEN: Database access.
ERRORS: None.

Извлечение полных строк с контекстом из Excel файлов.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Any

from loguru import logger
from openpyxl import load_workbook

from services.document_search.excel_utils import (
    xls_col_to_letter,
    letter_to_col_index,
    format_cell_display,
)

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


def extract_full_row_xls(
    file_path: Path,
    sheet_name: str,
    row_number: int,
    found_column: str,
    context_cols: int = 3,
) -> Dict[str, Any]:
    """Извлечение полной строки из .xls файла с контекстом."""
    if not XLRD_AVAILABLE:
        return {}
    
    try:
        workbook = xlrd.open_workbook(str(file_path))
    except Exception as error:
        logger.debug(f"Не удалось открыть .xls файл для извлечения полной строки: {error}")
        return {}
    
    try:
        sheet = workbook.sheet_by_name(sheet_name)
    except xlrd.XLRDError:
        logger.debug(f"Лист '{sheet_name}' не найден в файле {file_path}")
        return {}
    
    row_idx = row_number - 1
    if row_idx >= sheet.nrows:
        return {}
    
    found_col_idx = letter_to_col_index(found_column) - 1
    
    column_names = {}
    for header_row_idx in range(min(5, sheet.nrows)):
        for col_idx in range(min(sheet.ncols, 50)):
            cell = sheet.cell(header_row_idx, col_idx)
            if cell.value:
                col_letter = xls_col_to_letter(col_idx)
                if col_letter not in column_names or header_row_idx == 0:
                    column_names[col_letter] = str(cell.value).strip()
    
    full_row = []
    for col_idx in range(sheet.ncols):
        cell = sheet.cell(row_idx, col_idx)
        col_letter = xls_col_to_letter(col_idx)
        value = format_cell_display(cell.value) if cell.value else ""
        full_row.append({
            "column": col_letter,
            "column_name": column_names.get(col_letter, ""),
            "value": value
        })
    
    left_context = []
    right_context = []
    
    start_col = max(0, found_col_idx - context_cols)
    end_col = min(sheet.ncols, found_col_idx + context_cols + 1)
    
    for col_idx in range(start_col, found_col_idx):
        cell = sheet.cell(row_idx, col_idx)
        col_letter = xls_col_to_letter(col_idx)
        left_context.append({
            "column": col_letter,
            "column_name": column_names.get(col_letter, ""),
            "value": format_cell_display(cell.value) if cell.value else ""
        })
    
    for col_idx in range(found_col_idx + 1, end_col):
        cell = sheet.cell(row_idx, col_idx)
        col_letter = xls_col_to_letter(col_idx)
        right_context.append({
            "column": col_letter,
            "column_name": column_names.get(col_letter, ""),
            "value": format_cell_display(cell.value) if cell.value else ""
        })
    
    return {
        "full_row": full_row,
        "left_context": left_context,
        "right_context": right_context,
        "column_names": column_names
    }


def extract_full_row_xlsx(
    file_path: Path,
    sheet_name: str,
    row_number: int,
    found_column: str,
    context_cols: int = 3,
) -> Dict[str, Any]:
    """Извлечение полной строки из .xlsx файла с контекстом."""
    workbook = None
    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    except Exception as error:
        logger.debug(f"Не удалось открыть .xlsx файл для извлечения полной строки: {error}")
        return {}
    
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        logger.debug(f"Лист '{sheet_name}' не найден в файле {file_path}")
        return {}

    try:
        found_col_idx = letter_to_col_index(found_column)
        
        column_names = {}
        for header_row_idx in range(1, min(6, sheet.max_row + 1)):
            for col_idx in range(1, min(sheet.max_column + 1, 50)):
                cell = sheet.cell(row=header_row_idx, column=col_idx)
                if cell.value:
                    col_letter = cell.column_letter
                    if col_letter not in column_names or header_row_idx == 1:
                        column_names[col_letter] = str(cell.value).strip()
        
        full_row = []
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_number, column=col_idx)
            col_letter = cell.column_letter
            value = format_cell_display(cell.value) if cell.value is not None else ""
            full_row.append({
                "column": col_letter,
                "column_name": column_names.get(col_letter, ""),
                "value": value
            })
        
        left_context = []
        right_context = []
        
        start_col = max(1, found_col_idx - context_cols + 1)
        end_col = min(sheet.max_column + 1, found_col_idx + context_cols + 2)
        
        for col_idx in range(start_col, found_col_idx):
            cell = sheet.cell(row=row_number, column=col_idx)
            left_context.append({
                "column": cell.column_letter,
                "column_name": column_names.get(cell.column_letter, ""),
                "value": format_cell_display(cell.value) if cell.value is not None else ""
            })
        
        for col_idx in range(found_col_idx + 1, end_col):
            cell = sheet.cell(row=row_number, column=col_idx)
            right_context.append({
                "column": cell.column_letter,
                "column_name": column_names.get(cell.column_letter, ""),
                "value": format_cell_display(cell.value) if cell.value is not None else ""
            })
        
        return {
            "full_row": full_row,
            "left_context": left_context,
            "right_context": right_context,
            "column_names": column_names
        }
    finally:
        if workbook:
            workbook.close()

