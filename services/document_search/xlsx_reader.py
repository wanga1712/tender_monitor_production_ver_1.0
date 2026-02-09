"""
MODULE: services.document_search.xlsx_reader
RESPONSIBILITY: Read .xlsx files (using openpyxl).
ALLOWED: openpyxl, xlrd, pandas, file_format_detector, excel_utils, file_lock_handler, logging, core.exceptions.
FORBIDDEN: Database access.
ERRORS: DocumentSearchError.

Чтение .xlsx файлов через openpyxl.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Any, Iterable
import warnings

from loguru import logger
from openpyxl import load_workbook

from core.exceptions import DocumentSearchError
from services.document_search.file_format_detector import detect_file_format
from services.document_search.excel_utils import normalize_cell_value, format_cell_display
from services.document_search.file_lock_handler import handle_file_lock

# Подавляем предупреждения openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


def iter_xlsx_cells(
    file_path: Path,
    iter_xls_cells_func=None,
    iter_pandas_cells_func=None,
) -> Iterable[Dict[str, Any]]:
    """Итерация по .xlsx файлам через openpyxl."""
    try:
        # Используем handle_file_lock для обработки блокировок
        workbook = handle_file_lock(
            file_path,
            lambda: load_workbook(filename=file_path, read_only=True, data_only=True)
        )
    except Exception as error:
        error_msg = str(error).lower()
        detected_format = detect_file_format(file_path)
        
        if "file is not a zip file" in error_msg or "not a zip file" in error_msg or "bad magic number" in error_msg.lower():
            # Если файл определен как .xls или имеет расширение .xlsx, но не ZIP - пробуем через xlrd
            # Также пробуем, если формат unknown_xlsx (скорее всего это .xls с неправильным расширением)
            if detected_format in ('xls', 'unknown_xlsx') and XLRD_AVAILABLE and iter_xls_cells_func:
                logger.debug(
                    f"Файл {file_path.name} имеет расширение .xlsx, но на самом деле это .xls файл (определен как: {detected_format}). Пробую через xlrd."
                )
                try:
                    yield from iter_xls_cells_func(file_path, allow_fallback=False)
                    return
                except Exception as xlrd_error:
                    logger.debug(f"xlrd не смог открыть {file_path.name}: {xlrd_error}")
                    # Продолжаем попытки через другие методы
            elif detected_format in ('unknown_xlsx', 'unknown_xls', 'unknown'):
                logger.debug(
                    f"Файл {file_path.name} имеет неправильный формат (определен как: {detected_format}). "
                    f"Пробую открыть всеми доступными методами."
                )
                if XLRD_AVAILABLE and iter_xls_cells_func:
                    try:
                        yield from iter_xls_cells_func(file_path, allow_fallback=False)
                        return
                    except Exception:
                        pass
                if PANDAS_AVAILABLE and iter_pandas_cells_func:
                    # Проверяем доступность движков
                    available_engines = ['xlrd', 'openpyxl']
                    try:
                        import python_calamine
                        available_engines.append('calamine')
                    except ImportError:
                        logger.debug("python-calamine не установлен, пропускаем этот движок")
                    
                    for engine in available_engines:
                        try:
                            yield from iter_pandas_cells_func(file_path, engine=engine)
                            return
                        except Exception:
                            # Ошибки собираются внутри pandas_reader, просто продолжаем
                            continue
            else:
                logger.debug(
                    f"Файл {file_path.name} поврежден или имеет неправильный формат (определен как: {detected_format}). "
                    f"openpyxl не может открыть его как ZIP-архив. Пробую другие методы."
                )
        
        if "old .xls file format" in error_msg or "does not support" in error_msg:
            if PANDAS_AVAILABLE and iter_pandas_cells_func:
                logger.debug(
                    f"openpyxl не может открыть {file_path.name} (старый формат .xls). Пробую через pandas."
                )
                try:
                    yield from iter_pandas_cells_func(file_path, engine='xlrd')
                    return
                except Exception:
                    # Ошибка будет обработана ниже или в вызывающем коде
                    pass
        
        # Не логируем ERROR здесь - он будет залогирован в вызывающем коде (xls_reader или excel_parser)
        # Просто выбрасываем исключение для дальнейшей обработки
        raise DocumentSearchError(f"Поврежденный или неподдерживаемый XLSX файл: {error}") from error
    
    try:
        for sheet in workbook.worksheets:
            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                for cell in row:
                    if cell.value is None:
                        continue
                    text = normalize_cell_value(cell.value)
                    if text:
                        yield {
                            "text": text,
                            "display_text": format_cell_display(cell.value),
                            "sheet_name": sheet.title,
                            "row": row_idx,
                            "column": cell.column_letter,
                            "cell_address": f"{sheet.title}!{cell.coordinate}",
                        }
    finally:
        workbook.close()

