"""
MODULE: services.document_search.xlsx_row_extractor
RESPONSIBILITY: Extract specific row data from .xlsx sheets.
ALLOWED: openpyxl, xlsx_row_data_processor, excel_utils, error_logger, logging.
FORBIDDEN: Database access.
ERRORS: None.

Извлечение данных из строк .xlsx файлов.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Any

from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from services.error_logger import get_error_logger
from services.document_search.excel_utils import format_cell_display
from services.document_search.xlsx_row_data_processor import extract_row_data_from_values_xlsx

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


def extract_xlsx_row_data(
    file_path: Path,
    sheet_name: str,
    row_number: int,
    extract_xls_row_data_func=None,
) -> Dict[str, Any]:
    """Извлечение данных из .xlsx файла."""
    workbook = None
    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    except Exception as error:
        if XLRD_AVAILABLE and extract_xls_row_data_func:
            logger.warning(
                f"openpyxl не смог открыть {file_path} для извлечения данных строки (лист {sheet_name}): {error}. Пробую через xlrd."
            )
            try:
                return extract_xls_row_data_func(
                    file_path,
                    sheet_name,
                    row_number,
                )
            except Exception as xlrd_error:
                logger.error(f"xlrd также не смог открыть {file_path}: {xlrd_error}")
                return {}
        else:
            from services.document_search.file_format_detector import detect_file_format
            detected_format = detect_file_format(file_path)
            logger.error(f"Не удалось открыть документ {file_path}: {error}. xlrd недоступен для fallback.")
            get_error_logger().log_file_open_error(
                file_path=file_path,
                error_message=f"{error}. xlrd недоступен для fallback.",
                file_type="xlsx",
                detected_format=detected_format,
            )
            return {}
    
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        logger.warning(f"Лист '{sheet_name}' не найден в файле {file_path}")
        return {}

    try:
        row_values = []
        max_col = sheet.max_column
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_number, column=col_idx)
            row_values.append(format_cell_display(cell.value) if cell.value is not None else "")
        
        return extract_row_data_from_values_xlsx(row_values, max_col, sheet, get_column_letter)
    finally:
        if workbook:
            workbook.close()

