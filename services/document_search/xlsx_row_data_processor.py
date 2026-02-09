"""
MODULE: services.document_search.xlsx_row_data_processor
RESPONSIBILITY: Process extracted row values for .xlsx (helper for headers/columns).
ALLOWED: openpyxl.utils.
FORBIDDEN: IO operations.
ERRORS: None.

Обработка данных строк для .xlsx файлов.
"""

from __future__ import annotations

from typing import Dict, Any
from openpyxl.utils import get_column_letter


def extract_row_data_from_values_xlsx(
    row_values: list,
    max_col: int,
    sheet,
    get_column_letter_func,
) -> Dict[str, Any]:
    """Извлечение данных из значений строки для .xlsx."""
    column_headers = {}
    header_row = None
    
    for header_row_idx in range(1, min(6, sheet.max_row + 1)):
        for col_idx in range(1, min(max_col + 1, 20)):
            cell = sheet.cell(row=header_row_idx, column=col_idx)
            if cell.value:
                header_value = str(cell.value).strip().lower()
                if header_value == "количество" or (
                    header_value.startswith("количество") and "количество" not in column_headers
                ):
                    column_headers["количество"] = {"col": cell.column_letter, "idx": col_idx - 1}
                    if header_row is None:
                        header_row = header_row_idx
    
    if "количество" not in column_headers and max_col >= 4:
        column_headers["количество"] = {"col": "D", "idx": 3}
    
    cost_unit_col = None
    total_cost_col = None
    
    for header_row_idx in range(1, min(6, sheet.max_row + 1)):
        for col_idx in range(1, min(max_col + 1, 20)):
            cell = sheet.cell(row=header_row_idx, column=col_idx)
            if cell.value:
                header_value = str(cell.value).strip().lower()
                if ("стоимость единицы" in header_value or 
                    (header_value.startswith("стоимость") and "единицы" in header_value)):
                    if cost_unit_col is None:
                        cost_unit_col = {"col": cell.column_letter, "idx": col_idx - 1, "name": str(cell.value).strip()}
                if "общая стоимость" in header_value or header_value.startswith("общая стоимость"):
                    if total_cost_col is None:
                        total_cost_col = {"col": cell.column_letter, "idx": col_idx - 1, "name": str(cell.value).strip()}
    
    return build_result_dict_xlsx(row_values, column_headers, cost_unit_col, total_cost_col, max_col, get_column_letter_func)


def build_result_dict_xlsx(
    row_values: list,
    column_headers: dict,
    cost_unit_col: dict,
    total_cost_col: dict,
    max_col: int,
    get_column_letter_func,
) -> Dict[str, Any]:
    """Построение результирующего словаря для .xlsx."""
    result = {}
    if "количество" in column_headers:
        col_idx = column_headers["количество"]["idx"]
        if col_idx < len(row_values) and row_values[col_idx]:
            result["количество"] = {
                "value": row_values[col_idx],
                "column": column_headers["количество"]["col"],
                "name": "Количество"
            }
    
    if cost_unit_col:
        col_idx = cost_unit_col["idx"]
        if col_idx < len(row_values) and row_values[col_idx]:
            result["стоимость_единицы"] = {
                "value": row_values[col_idx],
                "column": cost_unit_col["col"],
                "name": cost_unit_col["name"]
            }
    elif max_col >= 5:
        for col_idx in [4, 5]:
            if col_idx < len(row_values) and row_values[col_idx]:
                result["стоимость_единицы"] = {
                    "value": row_values[col_idx],
                    "column": get_column_letter_func(col_idx + 1),
                    "name": "Стоимость единицы"
                }
                break
    
    if total_cost_col:
        col_idx = total_cost_col["idx"]
        if col_idx < len(row_values) and row_values[col_idx]:
            result["общая_стоимость"] = {
                "value": row_values[col_idx],
                "column": total_cost_col["col"],
                "name": total_cost_col["name"]
            }
    elif max_col >= 7:
        for col_idx in [6, 7, 8]:
            if col_idx < len(row_values) and row_values[col_idx]:
                result["общая_стоимость"] = {
                    "value": row_values[col_idx],
                    "column": get_column_letter_func(col_idx + 1),
                    "name": "Общая стоимость"
                }
                break
    
    return result

