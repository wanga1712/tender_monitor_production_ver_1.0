import os
from pathlib import Path
from typing import Optional, List
import pypdf
from pdf2image import convert_from_path
import pytesseract
from docx import Document
import openpyxl
from utils.logger_config import get_logger

logger = get_logger()

class BaseParser:
    def parse(self, file_path: Path) -> str:
        raise NotImplementedError

class PDFParser(BaseParser):
    def __init__(self, page_limit: int = 1):
        self.page_limit = page_limit

    def parse(self, file_path: Path) -> str:
        text = ""
        try:
            # Try to extract text first
            with open(file_path, 'rb') as f:
                reader = pypdf.PdfReader(f)
                num_pages = min(len(reader.pages), self.page_limit)
                
                for i in range(num_pages):
                    page_text = reader.pages[i].extract_text()
                    if page_text:
                        text += page_text + "\n"
            
            # If text is very short, assume scanned and try OCR
            if len(text.strip()) < 50:
                logger.info(f"PDF {file_path.name} seems scanned, trying OCR on first {self.page_limit} page(s)")
                try:
                    # Convert PDF to images
                    # thread_count=1 to save resources on weak server
                    images = convert_from_path(file_path, first_page=1, last_page=self.page_limit, thread_count=1)
                    
                    ocr_text = ""
                    for image in images:
                        ocr_text += pytesseract.image_to_string(image, lang='rus+eng') + "\n"
                    
                    if ocr_text.strip():
                        text = ocr_text
                except Exception as e:
                    logger.warning(f"OCR failed for {file_path.name}: {e}")
                    # If OCR fails (e.g. no poppler/tesseract), just return what we have
        
        except Exception as e:
            logger.error(f"Error parsing PDF {file_path}: {e}")
            
        return text

class WordParser(BaseParser):
    def parse(self, file_path: Path) -> str:
        text = ""
        try:
            doc = Document(file_path)
            # Limit? Word docs are usually text, so size isn't as big of an issue as OCR.
            # But we can limit paragraphs if needed. User only asked for page limit on PDF.
            for para in doc.paragraphs:
                text += para.text + "\n"
        except Exception as e:
            logger.error(f"Error parsing Word {file_path}: {e}")
        return text

class ExcelParser(BaseParser):
    def parse(self, file_path: Path) -> str:
        text = ""
        try:
            # read_only=True for performance
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            # Limit to first sheet? User said "в екселе могут быть листы".
            # Let's process all sheets but maybe limit rows?
            # User didn't explicitly ask for excel limit, only PDF.
            # But to be safe on weak server, maybe don't load everything into memory.
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    text += row_text + "\n"
        except Exception as e:
            logger.error(f"Error parsing Excel {file_path}: {e}")
        return text

class TextParser(BaseParser):
    def parse(self, file_path: Path) -> str:
        try:
            # Try utf-8 then cp1251
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except UnicodeDecodeError:
                with open(file_path, 'r', encoding='cp1251') as f:
                    return f.read()
        except Exception as e:
            logger.error(f"Error parsing Text {file_path}: {e}")
            return ""

class ParserFactory:
    def __init__(self, pdf_page_limit: int = 1):
        self.pdf_parser = PDFParser(page_limit=pdf_page_limit)
        self.word_parser = WordParser()
        self.excel_parser = ExcelParser()
        self.text_parser = TextParser()

    def get_parser(self, file_path: Path) -> Optional[BaseParser]:
        ext = file_path.suffix.lower()
        
        if ext == '.pdf':
            return self.pdf_parser
        elif ext in ['.docx', '.doc']:
            # python-docx only supports .docx. .doc requires external tools (like antiword or libreoffice)
            if ext == '.docx':
                return self.word_parser
            else:
                logger.warning(f".doc format not supported natively (needs .docx): {file_path}")
                return None
        elif ext in ['.xlsx', '.xlsm']:
            return self.excel_parser
        elif ext in ['.txt', '.csv', '.log', '.xml', '.json']:
            return self.text_parser
        else:
            return None
